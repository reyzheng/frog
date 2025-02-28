import openpyxl
import getopt, sys, json

myDB = dict()

def parse(inputFile):
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(inputFile)
    for sheet in dataframe.worksheets:
        #print(sheet)
        # Iterate the loop to read the cell values
        emptyRow = False
        for row in range(0, sheet.max_row):
            if row == 0:
                # skip first row
                continue
            key = ''
            year = 0
            colIdx = 0
            for col in sheet.iter_cols(1, sheet.max_column):
                #print(colIdx)
                if colIdx == 0:
                    if col[row].value is None:
                        #print("empty")
                        emptyRow = True
                        break
                elif colIdx == 1:
                    year = int(col[row].value)
                    #print("debug0 {}".format(year))
                elif colIdx == 2:
                    #print("debug1 " + col[row].value)
                    #key = col[row].value.strip().encode("utf-8").decode("utf-8")
                    key = col[row].value.strip()
                    #print(key)
                    #sys.exit(0)
                colIdx = colIdx + 1
            if emptyRow == True:
                break
            if key not in myDB:
                myDB[key] = dict()
                myDB[key]["namw"] = key
                myDB[key]["years"] = []
            myDB[key]['years'].append(year)
    #print(myDB)
    #print(json.dumps(myDB, indent=2))
    #sys.exit(1)

def exec(command):
    global myDB
    tokens = command.split(":")
    if tokens[0] == 'name':
        name = tokens[1]
        print("Query user {}".format(name))
        if name in myDB:
            years = list(set(myDB[name]['years']))
            print("\tyears: {}".format(years))
            print("\tcounts: {}".format(len(years)))
        else:
            print("\tUnknown user")
    elif tokens[0] == "cons":
        subcommands = tokens[1].split(",")
        year = int(subcommands[0])
        ite = int(subcommands[1])
        years = []
        for i in range(ite):
            years.append(year - i)
        print("Query sign {} times to {}({})".format(ite, year, years))
        total = 0
        for name in myDB:
            qualify = True
            for year in years:
                if year not in myDB[name]['years']:
                    qualify = False
            if qualify == True:
                total = total + 1
                print("\t{}, {}".format(name, myDB[name]['years']))
        print("\ttotal {}".format(total))
    elif tokens[0] == "conpercentage":
        year = int(tokens[1])
        prevyear = year - 1
        # both year, prevyear
        numerator = 0
        # prevyear only
        denominator = 0
        members = []
        for name in myDB:
            if year in myDB[name]['years'] and prevyear in myDB[name]['years']:
                numerator = numerator + 1
                members.append(name)
            if prevyear in myDB[name]['years']:
                denominator = denominator + 1
        print("Query continuous member from {} to {}".format(prevyear, year))
        print("\tcount {}".format(numerator))
        print("\tpercentag {}%".format((numerator * 100) / denominator))
        print("\tdetail {}".format(members))

def main():
    try:
        opts, args = getopt.getopt(sys.argv[1:], "c:f:o:v", ["command=", "file=", "output="])
    except getopt.GetoptError as err:
        print(err)  # 會印出像是 "option -a not recognized" 的訊息
        sys.exit(2)
    output = None
    verbose = False
    for o, a in opts:
        if o == "-v":
            verbose = True
        #elif o in ("-h", "--help"):
        #    sys.exit()
        elif o in ("-c", "--command"):
            command = a
        elif o in ("-f", "--file"):
            inputFile = a
        elif o in ("-o", "--output"):
            output = a
        else:
            assert False, "unhandled option"
    parse(inputFile)
    exec(command)

if __name__ == "__main__":
    main()